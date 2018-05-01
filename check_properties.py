from county_court_foreclose import CourtForeclosureSearch
from sheriff_foreclosures import SheriffForeclosureSearch
from datetime import date
from RealEstateListings import REListings
from Commercial import CommListings
import logging
import openpyxl
import excel_dict

# logger = logging.getLogger(__name__)


class SearchTargets:
    # one 'SearchTargets' object per file listing properties
    # can run various searches against it and save together

    def __init__(self, file, sheet):
        self.file = file + '.xlsx'
        self.sheet = sheet
        self.wb = openpyxl.load_workbook(self.file)
        self.address_list = []
        self.out_sheets = {}
        self.forc_s_list = []
        self.re_sale_list= []
        self.forc_ct_list = []
        self.comm_list = []

    def readfile(self):
        if not hasattr(self, 'address_list') or len(self.address_list) == 0:
        # if this instance of SearchTargets doesn't already have its list,
        # get it from the file
            sheet = self.wb.get_sheet_by_name(self.sheet)
            for row in range(1, sheet.max_row + 1):
                address = {"parcel": sheet['B' + str(row)].value, "address": sheet['C' + str(row)].value,
                           "owner": sheet['D' + str(row)].value}
                self.address_list.append(address)
                self.out_sheets['address list'] = self.address_list
        return self.address_list

    def write_results_sheet(self, result_list, result_type_descrip):
        date_string = date.today().isoformat()
        new_sheet_name = date_string + result_type_descrip
        # sheet = self.wb.create_sheet(new_sheet_name)
        self.out_sheets[new_sheet_name] = result_list
        return new_sheet_name, result_list


def targets_to_set(target_list, address_field):
    address_set = set()
    for entry in target_list:
        add = entry[address_field]
        if isinstance(add, str) and address_field == 'address':
            add = add.casefold()
            add = standardize(add)
        address_set.add(add)
        # print(address_set)
    return address_set


def standardize(entry):
    entry = entry.replace('w.', 'west')
    entry = entry.replace('e.', 'east')
    entry = entry.replace('n.', 'north')
    entry = entry.replace('s.', 'south')
    entry = entry.replace('2nd', 'second')
    entry = entry.replace('brook park', 'brookpark')
    # next only works if addresses in both sets of data always have
    # some designation ('street', 'road', etc) after street name
    # but it seems like they do
    entry = entry.rstrip('abcdefghijklmnopqrstuvwxyz.')
    entry = entry.strip()
    return entry


def check_foreclosure(city_name, q, address_field):
    logging.info('check against foreclosure case results')
    cfs_search = CourtForeclosureSearch(city_name, 'No')
    if len(q.forc_ct_list) > 0:
        s = q.forc_ct_list
    else:
        s = cfs_search.run_and_report()
        q.forc_ct_list = s
    q_set = targets_to_set(q.readfile(), address_field)
    s_set = targets_to_set(s, address_field)
    r_set = q_set.intersection(s_set)
    logging.debug(q_set)
    logging.debug(s_set)
    logging.debug(r_set)

    result_list = []
    for r in r_set:
        for res in s:
            test = res[address_field]
            if isinstance(test, str):
                # noinspection PyUnresolvedReferences
                test = test.casefold()
                test = standardize(test)
            if r == test:
                result_list.append(res)
    return result_list


def check_foreclosure_sale(city_name, start_date, end_date, q, address_field):
    logging.info('check against sheriff sale results')
    search = SheriffForeclosureSearch(city_name, 'No')
    if len(q.forc_s_list)> 0:
        s = q.forc_s_list
    else:
        s = search.run_and_report(start_date, end_date)
        q.forc_s_list = s
    q_set = targets_to_set(q.readfile(), address_field)
    s_set = targets_to_set(s, address_field)
    # print(q_set)
    # print(s_set)
    r_set = q_set.intersection(s_set)
    # print(r_set)
    logging.debug(q_set)
    logging.debug(s_set)

    result_list = []
    for r in r_set:
        for res in s:
            test = res[address_field]
            if isinstance(test, str):
                test = test.casefold()
                test = standardize(test)
            if r == test:
                result_list.append(res)
    return result_list


def check_classifieds(city_name, q, address_field):
    logging.info('check against RE listings')
    search = REListings(city_name, 'No')
    s = search.run_and_report()
    q_set = targets_to_set(q.readfile(), address_field)
    s_set = targets_to_set(s, address_field)
    r_set = q_set.intersection(s_set)
    logging.debug(q_set)
    logging.debug(s_set)

    result_list = []
    for r in r_set:
        if not r == '':
            for res in s:
                try:
                    test = res[address_field]
                except KeyError as err:
                    logging.error(str(err), exc_info=True)
                    test = ''
                finally:
                    if isinstance(test, str):
                    # noinspection PyUnresolvedReferences
                        test = test.casefold()
                        test = standardize(test)
                # exception handling needed for this one bc can be blank addresses in both q and s
                # but other check_ methods have no possible blank addresses on
                # their 's' sides
                        logging.debug(test)
                        if test == '':
                            logging.info('RE listing lacked address')
                            logging.info(str(res))
                        if test is not None and test != '' and r == test:
                            result_list.append(res)
    return result_list


def check_commercial(city_name, q, address_field):
    logging.info('check against RE commcl listings')
    search = CommListings(city_name, 'No')
    s = search.run_and_report()
    q_set = targets_to_set(q.readfile(), address_field)
    s_set = targets_to_set(s, address_field)
    r_set = q_set.intersection(s_set)
    logging.debug(q_set)
    logging.debug(s_set)

    result_list = []
    for r in r_set:
        if not r == '':
            for res in s:
                try:
                    test = res[address_field]
                except KeyError as err:
                    logging.error(str(err), exc_info=True)
                    test = ''
                finally:
                    if isinstance(test, str):
                    # noinspection PyUnresolvedReferences
                        test = test.casefold()
                        test = standardize(test)
                # exception handling needed for this one bc can be blank addresses in both q and s
                # but other check_ methods have no possible blank addresses on
                # their 's' sides
                        logging.debug(test)
                        if test == '':
                            logging.info('RE listing lacked address')
                            logging.info(str(res))
                        if test is not None and test != '' and r == test:
                            result_list.append(res)
    return result_list


# def check_all_write(city_name, file_name, start_date='', end_date='', sheet_name='Sheet1'):
#     if '.xlsx' not in file_name:
#         file_name = file_name + '.xlsx'
#     q = SearchTargets(file_name, sheet_name)
#     q.write_results(check_foreclosure_sale(city_name, start_date, end_date, q, 'address'), 'sale_addss')
#     q.write_results(check_foreclosure_sale(city_name, start_date, end_date, q, 'parcel'), 'sale_prcl')
#     q.write_results(check_foreclosure(city_name, q, 'address'), 'cases_adss')
#     q.write_results(check_foreclosure(city_name, q, 'parcel'), 'cases_prcl')
#     q.write_results(check_classifieds(city_name, q, 'address'), 'listing')
#     q.write_results(check_commercial(city_name, q, 'address'), 'comm_listing')

def check_all_write(city_name, file_name, start_date='', end_date='', sheet_name='Sheet1'):

    q = SearchTargets(file_name, sheet_name)
    q.write_results_sheet(check_commercial(city_name, q, 'address'), 'comm_listing')
    q.write_results_sheet(check_foreclosure_sale(city_name, start_date, end_date, q, 'address'), 'sale_addss')
    q.write_results_sheet(check_foreclosure_sale(city_name, start_date, end_date, q, 'parcel'), 'sale_prcl')
    q.write_results_sheet(check_foreclosure(city_name, q, 'address'), 'cases_adss')
    q.write_results_sheet(check_foreclosure(city_name, q, 'parcel'), 'cases_prcl')
    q.write_results_sheet(check_classifieds(city_name, q, 'address'), 'listing')

    file_name = file_name + '_update_' + date.today().isoformat()
    excel_dict.rewrite_multi_sheet(q.out_sheets, file_name)
    logging.info(file_name)


logging.basicConfig(level=logging.INFO)
check_all_write('Berea', 'Berea Properties- Beech Pearl and Second (004)')
check_all_write('Brook Park', 'Brook Park Properties (003)')
